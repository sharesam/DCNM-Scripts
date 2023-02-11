from openpyxl import load_workbook

wb_1 = load_workbook("get-hostname-serial.xlsx")
sheet_1 = wb_1['hostname-serial']
wb_2 = load_workbook("AZNET DC Hostnames.xlsx")
sheet_2 = wb_2['Sheet1']

for i in range(2, len(sheet_1['A']) + 1):
    sheet1_value = sheet_1[f'A{i}'].value
    for j in range(2, len(sheet_1['A']) + 1):
        sheet2_value = sheet_2[f'A{j}'].value
        if sheet1_value == sheet2_value:
            sheet_1[f'E{i}'].value = sheet_2[f'B{j}'].value
            wb_1.save("get-hostname-serial.xlsx")
            break
